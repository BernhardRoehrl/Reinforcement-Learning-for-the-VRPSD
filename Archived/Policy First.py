import random
import numpy as np
import os
import timeit
import openpyxl
from Apriori_Revisited import distance_matrix, apriori_list
from Distance_Matrix import capacity, demand_top, demand_bottom, depot_position, customer_spread

start_time = timeit.default_timer()
apriori_list = apriori_list
#random.seed(9001)
"""imported Apriori-route"""
data = {}
"""Stores the data for the problem."""
data['distance_matrix'] = distance_matrix

"""Create Lists and Parameters"""
customer_list = []
avg_distance = []
last_10k_avg_distance = []
test_list = []
capacity = capacity  # Vehicle Capacity Parameter
demand_bottom = demand_bottom # Smallest Value Demand can take
demand_top = demand_top  # Highest Value Demand can take

"""Additional Parameters for QLearning"""
action_refill = 0
action_serve = 1
action_space_size = [action_refill, action_serve]

num_episodes = 10000

# Set up Saving Process
index = 1
depot_prefix = depot_position  # 0M = Depot in der Mitte, 0E = Depot am Rand
solution_prefix = "P_"
customer_spread = customer_spread  # C = Clustered or R = Random Spread Customers
dataset_name = solution_prefix + depot_prefix + customer_spread + str(len(apriori_list) - 2) + "_C" + str(
    capacity) + "_D" + str(
    demand_bottom) + "-" + str(demand_top)
workbook_name = f"{solution_prefix}Results.xlsx"
file_name1 = dataset_name + '_output.txt'
file_name2 = dataset_name + '_q_table.txt'
file_path1 = "Experiments\\{}".format(dataset_name)
if not os.path.exists(file_path1):
    os.makedirs(file_path1)
try:
    workbook = openpyxl.load_workbook(workbook_name)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save(workbook_name)
if dataset_name in workbook.sheetnames:
    del workbook[dataset_name]
    worksheet = workbook.create_sheet(dataset_name)
else:
    worksheet = workbook.create_sheet(dataset_name)

worksheet.cell(row = 3, column=1, value = 'Customer')
worksheet.cell(row = 3, column=2, value = 'Action')
worksheet.cell(row = 3, column=3, value = 'Refill_Counter')
worksheet.cell(row = 3, column=4, value = 'Failure_Counter')
worksheet.cell(row = 3, column=5, value = 'Capacity at Decision')
worksheet.cell(row = 3, column=7, value = 'Per Thousand Episodes')
worksheet.cell(row = 3, column=8, value = 'Average Distance')
worksheet.cell(row = 3, column=9, value = 'Relative Change')
worksheet.cell(row = 3, column=11,value =  'Time for Computation in (s)')
worksheet.cell(row = 1, column=1, value = 'METHOD = ' + solution_prefix + ' DEPOT 0M: Mitte, 0E: Edge = ' + depot_prefix + ' CUSTOMER_SPREAD = '
                + customer_spread + str(len(apriori_list) - 2) + " CAPACITY = " + str(capacity) + " DEMANDS BETWEEN = "
                + str(demand_bottom) + "-" + str(demand_top))
worksheet.cell(row= 21, column=11, value= 'Result last 10k')
#workbook_results = xlsxwriter.Workbook('results')
#worksheet_results = workbook_results.add_worksheet(solution_prefix)

"""Function & Classes"""


def create_customer():
    """Creates Customer_List with random demands & position from Apriori_List"""
    customer_list.clear()
    for y in apriori_list:
        customer_x = Customer(random.randrange(demand_bottom, demand_top, 1), y)
        if customer_x.position == 0:
            customer_x.demand = 0
        customer_list.append(customer_x)
    return customer_list


class Service:
    """For the vehicle and tracking distance"""

    def __init__(self, position, capacity, distance, step_distance):
        self.capacity = capacity
        self.step = 0
        self.old_capacity = 0
        self.position = position
        self.distance = distance
        self.step_distance = step_distance
        self.refill_counter = 0
        self.failure_counter = 0
        self.current_episode_position = 0

    def reset_variables(self, capacity):
        """Resetting for outer loop"""
        create_customer()
        self.position = 0
        self.capacity = capacity
        self.distance = 0
        self.step_distance = 0
        self.step = 0
        self.old_capacity = 0
        self.refill_counter = 0
        self.failure_counter = 0
        return self, customer_list

    def position_update(self, next_customer):
        """Update vehicle position"""
        self.position = next_customer
        return self.position

    def distance_update(self, target):
        """Update total vehicle distance travelled"""
        self.distance += data['distance_matrix'][self.position][target]
        self.step_distance += data['distance_matrix'][self.position][target]
        return self.distance, self.step_distance

    def refill(self, current_step, customer):
        """The Agent chooses to refill. This means driving back and to the new customer"""
        self.distance += data['distance_matrix'][self.position][0]
        self.step_distance = data['distance_matrix'][self.position][0]
        self.position = 0
        self.capacity = capacity
        self.distance += data['distance_matrix'][self.position][customer]
        self.step_distance += data['distance_matrix'][self.position][customer]
        self.position_update(customer)
        self.capacity = self.capacity - customer_list[current_step].demand
        customer_list[current_step].demand = 0
        if episode > num_episodes-1000:
            self.refill_counter += 1
        return self.position, self.capacity, self.distance, customer_list[
            current_step].demand, self.step_distance, self.refill_counter

    def serve(self, current_step, customer):
        if self.capacity < customer_list[current_step].demand:
            """Failure 1: serve partially -> Go To depot -> Go back -> Fulfill serving"""
            """Hinfahren1 und customer partially bedienen"""
            if episode > num_episodes-1000:
                self.failure_counter += 1
            self.distance += data['distance_matrix'][self.position][customer]
            self.step_distance += data['distance_matrix'][self.position][customer]
            self.position_update(customer)
            customer_list[current_step].demand = customer_list[current_step].demand - self.capacity
            self.capacity = 0
            """Zurückfahren Vom partially bedienten Customer zum Depot"""
            self.distance += data['distance_matrix'][self.position][0]
            self.step_distance += data['distance_matrix'][self.position][0]
            self.position = 0
            self.capacity = capacity
            """Hinfahren2 Zurück zum Customer und rest bedienen"""
            self.distance += data['distance_matrix'][self.position][customer]
            self.step_distance += data['distance_matrix'][self.position][customer]
            self.position_update(customer)
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            return self.position, self.failure_counter, self.capacity, self.distance, customer_list[
                current_step].demand, self.step_distance
        else:
            """Vehicle can serve fully. regular workflow"""
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            self.distance_update(customer)
            self.position_update(customer)
            return customer_list[current_step].demand, self.capacity, self.position, self.distance, self.step_distance

    def execute_episode(self):
        self.step_distance = 0
        self.old_capacity = self.capacity
        if self.capacity > capacity*0.25:
            """Agent chooses action 1: Vehicle must serve"""
            action = 1
            self.serve(self.step, apriori_list[self.step])
        else:
            """Agent chooses action 2: refilling"""
            action = 0
            self.refill(self.step, apriori_list[self.step])
        self.step = self.step + 1
        if episode == num_episodes-1:
            self.write_final_episode(action)
        return self

    def post_episode_calculation(self):
        """After each episode this updates the exploration rate according to e-greedy and
        adds the data for rewards and distances to the lists for later calculations"""
        avg_distance.append(self.distance)
        return avg_distance, self

    def write_final_episode(self, action):
        """Function that writes the best routing found after training to text and excel"""
        file_path = os.path.join(file_path1, file_name1)
        worksheet.cell(row = self.step+3, column = 1, value = x)
        worksheet.cell(row = self.step+3, column = 2, value = action)
        worksheet.cell(row = self.step+3, column = 3, value = self.refill_counter)
        worksheet.cell(row = self.step+3, column = 4, value = self.failure_counter)
        worksheet.cell(row = self.step+3, column = 5, value = self.old_capacity)
        with open(file_path, 'a') as outfile:
            outfile.write("Customer: " + str(x) + " Action: " + str(action) + " Refill_Counter: " +
                          str(self.refill_counter) + " Failure_counter: " + str(self.failure_counter) + "\n")


class Customer:
    """Customers with demand and position"""

    def __init__(self, demand, position):
        self.demand = demand
        self.position = position

    def set_demand(self, demand):
        """Change Customer Demand"""
        self.demand = demand


def print_final(file_path1, file_name1, row_position):
    """Function that incorporates all output related information for further useage"""
    avg_distances_per_thousand_episodes = np.split(np.array(avg_distance), num_episodes / 1000)
    slice_index = max(0, len(avg_distance) -10000)
    last_10k_distances = avg_distance[slice_index:]
    last_10k_avg_distances = np.mean(last_10k_distances)
    count = 1000
    rev_value = None
    # Write to log-file
    file_path = os.path.join(file_path1, file_name1)
    with open(file_path, 'a') as outfile:
        print("\nAverage Distance per thousand episodes\n")
        outfile.write("\nAverage Distance per thousand episodes\n")
        for d in avg_distances_per_thousand_episodes:
            print(count, ":", str(sum(d / 1000)))
            outfile.write(str(count) + ":" + str(sum(d / 1000)) + "\n")
            # Write to Excel
            worksheet.cell(row=row_position + 3 , column = 7, value = count)
            worksheet.cell(row=row_position + 3 , column = 8, value = sum(d / 1000))
            if rev_value is None:
                ref_value = worksheet.cell(row=4, column =8).value
            worksheet.cell(row=row_position +3 , column = 9, value = sum(d /1000)/ref_value)
            count += 1000
            row_position += 1
        worksheet.cell(row=22, column=11, value=last_10k_avg_distances)
        # logfile and output print for q_Table and exploration_counter
    worksheet.cell(row = 4, column = 11, value = elapsed_time)   # Computational Time to Excel
    # Processing the 3d Numpy Array: q_table for Excel
    workbook.save(workbook_name)

if __name__ == "__main__":
    vehicle = Service(0, capacity, 0, 0)   # Init the vehicle
    for episode in range(num_episodes):   # Outer Loop for 16000 Episodes
        vehicle.reset_variables(capacity)   # Resetting the variables that got changed in inner loop
        for x in apriori_list:   # Inner Loop to let the vehicle go through all customers
            vehicle.execute_episode()   # Main function executing the service
        vehicle.post_episode_calculation()   # Capture and prepare calculations after every episode
    elapsed_time = timeit.default_timer() - start_time  # Computational Time
    print_final(file_path1, file_name1, index)   # Save results for later use


import pandas as pd
import openpyxl
import os

# Excels file set up
input_filename1 = 'P_Results.xlsx'
input_filename2 = 'RL_Results.xlsx'
output_filename = 'Excel_Analysis.xlsx'
workbook_name = f"{'Excel_Analysis.xlsx'}"

def save_to_excel(df, filename, sheet_name):
    book = openpyxl.load_workbook(filename)
    if sheet_name in book.sheetnames:  # check if sheet_name already exists
        if book[sheet_name].max_row > 1:  # check if sheet contains data (not empty)
            bouncer = input("Do you want to overwrite the sheet?")
            while bouncer != "no":
                if bouncer.lower() == "yes":
                    bouncer = input("Are you sure you want to overwrite the Data? (Yes/No): ")
                    if bouncer.lower() == "yes":
                        except yes:
                    elif bouncer.lower() == "no":
                        raise Exception("Terminated: User is not sure to overwrite")
                    else:
                        print("Invalid Answer provided by User")
                        bouncer = input("Do you want to overwrite the sheet?")
                        continue
                elif bouncer.lower() == "no":
                    raise Exception("Invalid Input or Terminated by User")
                else:
                    print("Invalid Answer provided by User")
                    bouncer = input("Do you want to overwrite the sheet?")
                    continue
            raise Exception("Terminated: User doesnt want to overwrite Data")
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        if sheet_name in writer.sheets:
            startrow = writer.sheets[sheet_name].max_row
            df.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name=sheet_name)
        else:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    else:  # if file doesn't exist, create a new one
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)


def extract_data(input_filename):
    """Extract Required Data from Excel Files"""

    # List to Store data in
    data_list = []
    #  Read Data from Excel File
    xls = pd.ExcelFile(input_filename)

    # Go Through every Tab (Sheet) in Excel File
    for sheet_name in xls.sheet_names:
        # Read the current Tab
        df = xls.parse(sheet_name)

        # Targeted Cells
        value1 = sheet_name  # Name aus Sheet da Sheet = Instance
        value2 = df.iloc[20, 10]  # Wert aus Zelle K22

        # Füge die Werte der Datenliste hinzu
        data_list.append([value1, value2])

    return data_list

# Call function to extract data for both excels
data1 = extract_data(input_filename1)
data2 = extract_data(input_filename2)

df_output1 = pd.DataFrame(data1, columns=['Instance', 'SP Result'])
df_output2 = pd.DataFrame(data2, columns=['Instance', 'RL Result'])

# Merge Instances to get 1 Column only IF the names are identical to prevent mistakes
df_output = pd.merge(df_output1, df_output2, on='Instance')

save_to_excel(df=df_output, filename=output_filename, sheet_name='Results')



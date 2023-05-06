import random
import numpy as np
import os
import pandas as pd
import timeit
import openpyxl

start_time = timeit.default_timer()
apriori_list = [0, 3, 4, 1, 2, 11, 12, 13, 14, 15, 16, 5, 6, 7, 8, 9, 10, 0]
random.seed(9001)
"""imported Apriori-route"""
data = {}
"""Stores the data for the problem."""
data['distance_matrix'] = [
    [
        0, 548, 776, 696, 582, 274, 502, 194, 308, 194, 536, 502, 388, 354,
        468, 776, 662
    ],
    [
        548, 0, 684, 308, 194, 502, 730, 354, 696, 742, 1084, 594, 480, 674,
        1016, 868, 1210
    ],
    [
        776, 684, 0, 992, 878, 502, 274, 810, 468, 742, 400, 1278, 1164,
        1130, 788, 1552, 754
    ],
    [
        696, 308, 992, 0, 114, 650, 878, 502, 844, 890, 1232, 514, 628, 822,
        1164, 560, 1358
    ],
    [
        582, 194, 878, 114, 0, 536, 764, 388, 730, 776, 1118, 400, 514, 708,
        1050, 674, 1244
    ],
    [
        274, 502, 502, 650, 536, 0, 228, 308, 194, 240, 582, 776, 662, 628,
        514, 1050, 708
    ],
    [
        502, 730, 274, 878, 764, 228, 0, 536, 194, 468, 354, 1004, 890, 856,
        514, 1278, 480
    ],
    [
        194, 354, 810, 502, 388, 308, 536, 0, 342, 388, 730, 468, 354, 320,
        662, 742, 856
    ],
    [
        308, 696, 468, 844, 730, 194, 194, 342, 0, 274, 388, 810, 696, 662,
        320, 1084, 514
    ],
    [
        194, 742, 742, 890, 776, 240, 468, 388, 274, 0, 342, 536, 422, 388,
        274, 810, 468
    ],
    [
        536, 1084, 400, 1232, 1118, 582, 354, 730, 388, 342, 0, 878, 764,
        730, 388, 1152, 354
    ],
    [
        502, 594, 1278, 514, 400, 776, 1004, 468, 810, 536, 878, 0, 114,
        308, 650, 274, 844
    ],
    [
        388, 480, 1164, 628, 514, 662, 890, 354, 696, 422, 764, 114, 0, 194,
        536, 388, 730
    ],
    [
        354, 674, 1130, 822, 708, 628, 856, 320, 662, 388, 730, 308, 194, 0,
        342, 422, 536
    ],
    [
        468, 1016, 788, 1164, 1050, 514, 514, 662, 320, 274, 388, 650, 536,
        342, 0, 764, 194
    ],
    [
        776, 868, 1552, 560, 674, 1050, 1278, 742, 1084, 810, 1152, 274,
        388, 422, 764, 0, 798
    ],
    [
        662, 1210, 754, 1358, 1244, 708, 480, 856, 514, 468, 354, 844, 730,
        536, 194, 798, 0
    ],
]

"""Create Lists and Parameters"""
customer_list = []
avg_distance = []
test_list = []
capacity = 50  # Vehicle Capacity Parameter
demand_bottom = 10  # Smallest Value Demand can take
demand_top = 50  # Highest Value Demand can take

"""Additional Parameters for QLearning"""
action_refill = 0
action_serve = 1
action_space_size = [action_refill, action_serve]

num_episodes = 16000
rewards_all_episodes = []  # List of rewards

learning_rate = 0.04  # Learning rate
discount_rate = 0.75  # Discounting rate

# Exploration parameters
exploration_rate = 1.0  # Exploration rate
max_exploration_rate = 1.0  # Exploration probability at start
min_exploration_rate = 0.001  # Minimum exploration probability
exploration_decay_rate = 0.05  # Exponential decay rate for exploration prob
exploration_counter = 0
list_q_table = [300]
q_table = np.zeros((capacity + 1, len(apriori_list) + 1, len(action_space_size)))  # init the q_table

# Set up Saving Process
index = 1
depot_prefix = "0M_"  # 0M = Depot in der Mitte, 0E = Depot am Rand
solution_prefix = "RL_"
customer_spread = "C"  # C = Clustered or R = Random Spread Customers
dataset_name = solution_prefix + depot_prefix + customer_spread + str(len(apriori_list) - 2) + "_C" + str(
    capacity) + "_D" + str(
    demand_bottom) + "-" + str(demand_top)
workbook_name = f"{solution_prefix}Results.xlsx"
file_name1 = dataset_name + '_output.txt'
file_name2 = dataset_name + '_q_table.txt'
file_path1 = "Experiments\\{}".format(dataset_name)
if not os.path.exists(file_path1):
    os.makedirs(file_path1)
try:
    workbook = openpyxl.load_workbook(workbook_name)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save(workbook_name)
if dataset_name in workbook.sheetnames:
    worksheet = workbook[dataset_name]
else:
    worksheet = workbook.create_sheet(dataset_name)

worksheet.cell(row = 3, column=1, value = 'Customer')
worksheet.cell(row = 3, column=2, value = 'Action')
worksheet.cell(row = 3, column=3, value = 'Refill_Counter')
worksheet.cell(row = 3, column=4, value = 'Failure_Counter')
worksheet.cell(row = 3, column=7, value = 'Per Thousand Episodes')
worksheet.cell(row = 3, column=8, value = 'Average Distance')
worksheet.cell(row = 3, column=11,value =  'Time for Computation in (s)')
worksheet.cell(row = 1, column=1, value = 'METHOD = ' + solution_prefix + ' DEPOT 0M: Mitte, 0E: Edge = ' + depot_prefix + ' CUSTOMER_SPREAD = '
                + customer_spread + str(len(apriori_list) - 2) + " CAPACITY = " + str(capacity) + " DEMANDS BETWEEN = "
                + str(demand_bottom) + "-" + str(demand_top))
#workbook_results = xlsxwriter.Workbook('results')
#worksheet_results = workbook_results.add_worksheet(solution_prefix)

"""Function & Classes"""


def create_customer():
    """Creates Customer_List with random demands & position from Apriori_List"""
    customer_list.clear()
    for y in apriori_list:
        customer_x = Customer(random.randrange(demand_bottom, demand_top, 1), y)
        if customer_x.position == 0:
            customer_x.demand = 0
        customer_list.append(customer_x)
    return customer_list


class Service:
    """For the vehicle and tracking distance"""

    def __init__(self, position, capacity, distance, step_distance):
        self.exploration_rate = exploration_rate
        self.capacity = capacity
        self.step = 0
        self.old_capacity = 0
        self.rewards_current_episodes = 0
        self.position = position
        self.distance = distance
        self.step_distance = step_distance
        self.refill_counter = 0
        self.failure_counter = 0
        self.exploration_counter = 0
        self.current_episode_position = 0

    def reset_variables(self, capacity):
        """Resetting for outer loop"""
        create_customer()
        self.position = 0
        self.capacity = capacity
        self.distance = 0
        self.step_distance = 0
        self.step = 0
        self.old_capacity = 0
        self.rewards_current_episodes = 0
        self.refill_counter = 0
        self.failure_counter = 0
        return self, customer_list

    def position_update(self, next_customer):
        """Update vehicle position"""
        self.position = next_customer
        return self.position

    def distance_update(self, target):
        """Update total vehicle distance travelled"""
        self.distance += data['distance_matrix'][self.position][target]
        self.step_distance += data['distance_matrix'][self.position][target]
        return self.distance, self.step_distance

    def refill(self, current_step, customer):
        """The Agent chooses to refill. This means driving back and to the new customer"""
        self.distance += data['distance_matrix'][self.position][0]
        self.step_distance = data['distance_matrix'][self.position][0]
        self.position = 0
        self.capacity = capacity
        self.distance += data['distance_matrix'][self.position][customer]
        self.step_distance += data['distance_matrix'][self.position][customer]
        self.position_update(customer)
        self.capacity = self.capacity - customer_list[current_step].demand
        customer_list[current_step].demand = 0
        if episode > 15000:
            self.refill_counter += 1
        return self.position, self.capacity, self.distance, customer_list[
            current_step].demand, self.step_distance, self.refill_counter

    def serve(self, current_step, customer):
        if self.capacity < customer_list[current_step].demand:
            """Failure 1: serve partially -> Go To depot -> Go back -> Fulfill serving"""
            """Hinfahren1 und customer partially bedienen"""
            if episode > 15000:
                self.failure_counter += 1
            self.distance += data['distance_matrix'][self.position][customer]
            self.step_distance += data['distance_matrix'][self.position][customer]
            self.position_update(customer)
            customer_list[current_step].demand = customer_list[current_step].demand - self.capacity
            self.capacity = 0
            """Zurückfahren Vom partially bedienten Customer zum Depot"""
            self.distance += data['distance_matrix'][self.position][0]
            self.step_distance += data['distance_matrix'][self.position][0]
            self.position = 0
            self.capacity = capacity
            """Hinfahren2 Zurück zum Customer und rest bedienen"""
            self.distance += data['distance_matrix'][self.position][customer]
            self.step_distance += data['distance_matrix'][self.position][customer]
            self.position_update(customer)
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            return self.position, self.failure_counter, self.capacity, self.distance, customer_list[
                current_step].demand, self.step_distance
        else:
            """Vehicle can serve fully. regular workflow"""
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            self.distance_update(customer)
            self.position_update(customer)
            return customer_list[current_step].demand, self.capacity, self.position, self.distance, self.step_distance

    def execute_episode(self):
        self.step_distance = 0
        self.old_capacity = self.capacity
        exploration_rate_threshold = random.uniform(0, 1)
        if exploration_rate_threshold < self.exploration_rate or episode < 15000:
            action = random.choice(action_space_size)
            if action == 1:
                """Agent chooses action 1: Vehicle must serve"""
                self.serve(self.step, apriori_list[self.step])
                self.exploration_counter += 1
            else:
                """Agent chooses action 2: refilling"""
                self.refill(self.step, apriori_list[self.step])
                self.exploration_counter += 1
        else:
            action = np.argmin(q_table[self.old_capacity, self.step, :])
            if action == 1:
                self.serve(self.step, apriori_list[self.step])
            elif action == 0:
                self.refill(self.step, apriori_list[self.step])
        reward = self.step_distance
        if episode < 15000:
            old_value = q_table[self.old_capacity, apriori_list[self.step], action]
            if self.step+1 < len(apriori_list):
                best_expected_value = np.min(q_table[self.capacity, apriori_list[self.step+1], :])
            else:
                best_expected_value = 0
            bellman_term = (reward + discount_rate * best_expected_value - old_value)
            q_table[self.old_capacity, self.current_episode_position, action] = old_value + learning_rate * bellman_term
        self.rewards_current_episodes += reward
        self.step = self.step + 1
        if episode == 15999:
            self.write_final_episode(action)
        return self, q_table

    def post_episode_calculation(self):
        """After each episode this updates the exploration rate according to e-greedy and
        adds the data for rewards and distances to the lists for later calculations"""
        avg_distance.append(self.distance)
        self.exploration_rate = min_exploration_rate + (max_exploration_rate - min_exploration_rate) * np.exp(
            -exploration_decay_rate * episode)
        rewards_all_episodes.append(self.rewards_current_episodes)
        for divider in list_q_table:
            if divider == episode:
                print(divider, "The Simulation explored: ", self.exploration_counter)
        return avg_distance, self, rewards_all_episodes,

    def write_final_episode(self, action):
        """Function that writes the best routing found after training to text and excel"""
        file_path = os.path.join(file_path1, file_name1)
        worksheet.cell(row = self.step+3, column = 1, value = x)
        worksheet.cell(row = self.step+3, column = 2, value = action)
        worksheet.cell(row = self.step+3, column = 3, value = self.refill_counter)
        worksheet.cell(row = self.step+3, column = 4, value = self.failure_counter)
        worksheet.cell(row = self.step+3, column = 5, value = self.old_capacity)
        with open(file_path, 'a') as outfile:
            outfile.write("Customer: " + str(x) + " Action: " + str(action) + " Refill_Counter: " +
                          str(self.refill_counter) + " Failure_counter: " + str(self.failure_counter) + "\n")


class Customer:
    """Customers with demand and position"""

    def __init__(self, demand, position):
        self.demand = demand
        self.position = position

    def set_demand(self, demand):
        """Change Customer Demand"""
        self.demand = demand


def print_final(file_path1, file_name1, data, Counter2):
    """Function that incorporates all output related information for further useage"""
    avg_distances_per_thousand_episodes = np.split(np.array(avg_distance), num_episodes / 1000)
    count = 1000
    # Write to log-file
    file_path = os.path.join(file_path1, file_name1)
    with open(file_path, 'a') as outfile:
        print("\nAverage Distance per thousand episodes\n")
        outfile.write("\nAverage Distance per thousand episodes\n")
        for d in avg_distances_per_thousand_episodes:
            print(count, ":", str(sum(d / 1000)))
            outfile.write(str(count) + ":" + str(sum(d / 1000)) + "\n")
            # Write to Excel
            worksheet.cell(row= Counter2+1 + 2, column = 7, value = count)
            worksheet.cell(row= Counter2+1 + 2, column = 8, value = sum(d / 1000))
            count += 1000
            Counter2 += 1
        # logfile and output print for q_Table and exploration_counter
        print("The Simulation explored: ", vehicle.exploration_counter)
        outfile.write("Explored: " + str(vehicle.exploration_counter) + "\n")
        print(q_table)
        outfile.write('\nThe Q_Table {0}\n'.format(data.shape))  # Header 1
        outfile.write('\nAction 0 : Action 1\n')  # Header 2
        for i, data_slice in enumerate(data):
            np.savetxt(outfile, data_slice, fmt='%-7.2f', delimiter=":")
            outfile.write('# Capacity: {0}\n'.format(i + 1))
    elapsed_time = timeit.default_timer() - start_time   # Computational Time
    worksheet.cell(row = 4, column = 11, value = elapsed_time)   # Computational Time to Excel
    # Processing the 3d Numpy Array: q_table for Excel
    names = ['x', 'y', 'z']
    header_names = pd.MultiIndex.from_product([range(s)for s in q_table.shape], names=names)
    df = pd.DataFrame({'q_table': q_table.flatten()}, index=header_names)['q_table']
    df = df.unstack(level='z').swaplevel().sort_index()
    df.columns = ['Refill', 'Serve']
    df.index.names = ['Customer', 'Capacity']
    worksheet.cell(row = 5, column = 16, value = 'Q-Table')  # Header for q_table
    workbook.save(workbook_name)
    writer = pd.ExcelWriter(workbook_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    with writer as writer:
        df.to_excel(writer, sheet_name=dataset_name, startrow=5, startcol=15)

if __name__ == "__main__":
    vehicle = Service(0, capacity, 0, 0)   # Init the vehicle
    for episode in range(num_episodes):   # Outer Loop for 16000 Episodes
        vehicle.reset_variables(capacity)   # Resetting the variables that got changed in inner loop
        for x in apriori_list:   # Inner Loop to let the vehicle go through all customers
            vehicle.execute_episode()   # Main function executing the service
        vehicle.post_episode_calculation()   # Capture and prepare calculations after every episode
    print_final(file_path1, file_name1, q_table, index)   # Save results for later use


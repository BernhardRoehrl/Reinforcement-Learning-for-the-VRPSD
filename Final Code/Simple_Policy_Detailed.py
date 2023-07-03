import random
import numpy as np
import os
import timeit
import openpyxl
from Apriori import Apriori
from Instance_Prerequisites import Instance

"""Import Data"""
start_time = timeit.default_timer()  # Set start tim
"""Create Lists and Parameters"""
customer_list = []  # Init customer list
avg_distance = []  # Init avg_distance
last_10k_avg_distance = []  # Init List for Results out of Benchmarking
num_episodes = 10000  # Amounts of Episodes the Outer Loop will do. 10k Episodes for Benchmarking

def LoadIn_instance(instance):
    apriori_list = Apriori.create_apriori_list(instance)
    data = {}  # Stores the data for the problem
    data['distance_matrix'] = instance.distance_matrix  # Import distance_matrix
    capacity = instance.capacity  # Vehicle Capacity Parameter
    demand_bottom = instance.demand_bottom  # Smallest Value Demand customer can have
    demand_top = instance.demand_top
    demand_values = [demand_top, demand_bottom]  # Highest Value Demand customer can have
    demand_mean = np.mean(demand_values)  # Average Demand of customers used to decide
    return data, demand_bottom, demand_top, capacity, apriori_list, demand_mean

def saving_process(instance):
    index = 1  # for looping later on
    solution_prefix = "P_"  # Indication of method
    dataset_name = instance.instance_name  # Dynamic Instance Naming
    workbook_name = f"{solution_prefix}Results.xlsx"
    try:  # Excel File Handling
        workbook = openpyxl.load_workbook(workbook_name)
    except FileNotFoundError:  # if Excel does not exist, create one
        workbook = openpyxl.Workbook()
        workbook.save(workbook_name)
    if dataset_name in workbook.sheetnames:  # if Tab in Excel exists delete and recreate
        del workbook[dataset_name]
        worksheet = workbook.create_sheet(dataset_name)
    else:  # Create if missing
        worksheet = workbook.create_sheet(dataset_name)
    # Write Headers and existing values
    worksheet.cell(row=1, column=1, value=dataset_name)
    worksheet.cell(row=3, column=1, value='Customer')
    worksheet.cell(row=3, column=2, value='Action')
    worksheet.cell(row=3, column=3, value='Refill_Counter')
    worksheet.cell(row=3, column=4, value='Failure_Counter')
    worksheet.cell(row=3, column=5, value='Capacity at Decision')
    worksheet.cell(row=3, column=7, value='Per Thousand Episodes')
    worksheet.cell(row=3, column=8, value='Average Distance')
    worksheet.cell(row=3, column=9, value='Relative Change')
    worksheet.cell(row=3, column=11, value='Time for Computation in (s)')
    worksheet.cell(row=21, column=11, value='Result last 10k')
    worksheet.cell(row=24, column=11, value='Failures')
    return workbook, worksheet, dataset_name, workbook_name

def create_customer(instance, apriori_list):
    """Creates Customer_List with random demands & position from Apriori_List"""
    customer_list.clear()
    for y in apriori_list:
        customer_x = Customer(random.randrange(instance.demand_bottom, instance.demand_top, 1), y)
        if customer_x.position == 0:
            customer_x.demand = 0
        customer_list.append(customer_x)
    if instance.capacity < instance.demand_top:
        #  Forbidden due to how customer demand and route failure are calculated
        raise Exception("!!!Demand > Capacity!!!")
    return customer_list


class Service:
    """For the vehicle and tracking distance"""

    def __init__(self, position, capacity, distance, step_distance):
        self.capacity = capacity
        self.step = 0
        self.old_capacity = 0  # Needed for Bellman Equation Calculations
        self.position = position
        self.distance = distance  # Track total distance throughout one episode
        self.step_distance = step_distance  # Track total distance for one customer
        self.refill_counter = 0  # For last episode tracking
        self.failure_counter = 0  # For last episode tracking
        self.failure_result = 0  # For whole Benchmarking Process

    def reset_variables(self, capacity, instance, apriori_list):
        """Resetting for outer loop"""
        create_customer(instance, apriori_list)
        self.position = 0
        self.capacity = capacity
        self.distance = 0
        self.step_distance = 0
        self.step = 0
        self.old_capacity = 0
        self.refill_counter = 0
        self.failure_counter = 0
        return self, customer_list

    def position_update(self, next_customer):
        """Update vehicle position"""
        self.position = next_customer
        return self.position

    def distance_update(self, target, data):
        """Update total vehicle distance travelled"""
        self.distance += data['distance_matrix'][self.position][target]  # Update total Episode Distance
        self.step_distance += data['distance_matrix'][self.position][target]  # Update Customer/Step Specific Distance
        return self.distance, self.step_distance

    def refill(self, current_step, customer, data, capacity,):
        """The Agent chooses to refill. This means driving back and to the new customer"""
        """Move Vehicle from current position to the depot and refill"""
        self.distance_update(0, data)
        self.position_update(0)
        self.capacity = capacity  # Refilling the Capacity
        """Move Vehicle from depot to next customer and serve"""
        self.distance_update(customer, data)
        self.position_update(customer)
        self.capacity = self.capacity - customer_list[current_step].demand  # Update capacity
        customer_list[current_step].demand = 0  # Set customer demand to 0
        self.refill_counter += 1  # For Benchmarking statistics
        return self.position, self.capacity, self.distance, customer_list[
            current_step].demand, self.step_distance, self.refill_counter

    def serve(self, current_step, customer, data, capacity,):
        """Function handling both scenarios for serving. Failure or successful 1st time-servings"""
        if self.capacity < customer_list[current_step].demand:
            """Failure: serve partially -> Go To depot -> Go back -> Fulfill serving"""
            # 1st Distance: Serve customer partially
            self.failure_counter += 1  # For Last Episode Overview
            self.failure_result += 1  # For Benchmarking statistics
            self.distance_update(customer, data)
            self.position_update(customer)
            customer_list[current_step].demand = customer_list[current_step].demand - self.capacity
            self.capacity = 0
            # 2nd Distance: To Depot from location of partially served customer
            self.distance_update(0, data)
            self.position_update(0)
            self.capacity = capacity
            # 3rd Distance: Back To Customer from Depot serving left over demand
            self.distance_update(customer, data)
            self.position_update(customer)
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            return self.position, self.failure_counter, self.capacity, self.distance, customer_list[
                current_step].demand, self.step_distance, self.failure_result
        else:
            """Vehicle can serve fully. regular workflow"""
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            self.distance_update(customer, data)
            self.position_update(customer)
            return customer_list[current_step].demand, self.capacity, self.position, self.distance, self.step_distance

    def execute_episode(self, demand_mean, apriori_list, data, capacity, worksheet, episode, x):
        """Core function executing the service for one step/customer/decision epoch"""
        self.step_distance = 0  # Reset step distance at the beginning of each
        self.old_capacity = self.capacity  # Update Capacities
        if self.capacity > demand_mean:  # Rule for Serving
            """Policy chooses action 1: Vehicle must serve"""
            action = 1
            self.serve(self.step, apriori_list[self.step], data, capacity)
        else:  # Rule for refilling
            """Agent chooses action 2: refilling"""
            action = 0
            self.refill(self.step, apriori_list[self.step], data, capacity)
        self.step = self.step + 1  # Adjustments for next episode
        if episode == num_episodes - 1:  # Write last Episode to Excel for quick overview
            self.write_final_episode(action, worksheet, x)
        return self

    def post_episode_calculation(self):
        """After each episode this updates the exploration rate according to e-greedy and
        adds the data for rewards and distances to the lists for later calculations"""
        avg_distance.append(self.distance)
        return avg_distance, self

    def write_final_episode(self, action, worksheet, x):
        """Function that writes the best routing found after training to text and excel"""
        worksheet.cell(row=self.step + 3, column=1, value=x)  # x = which customer
        worksheet.cell(row=self.step + 3, column=2, value=action)
        worksheet.cell(row=self.step + 3, column=3, value=self.refill_counter)
        worksheet.cell(row=self.step + 3, column=4, value=self.failure_counter)
        worksheet.cell(row=self.step + 3, column=5, value=self.old_capacity)


class Customer:
    """Customers with demand and position"""

    def __init__(self, demand, position):
        self.demand = demand
        self.position = position

    def set_demand(self, demand):
        """Change Customer Demand"""
        self.demand = demand


def print_final(row_position, worksheet, workbook, workbook_name, vehicle):
    """Function that incorporates all output related information for further usage, give index for row_position"""
    """Calculate All Kinds of Distances for Evaluation out of Lists"""
    avg_distances_per_thousand_episodes = np.split(np.array(avg_distance), num_episodes / 1000)
    slice_index = max(0, len(avg_distance) - 10000)
    last_10k_distances = avg_distance[slice_index:]
    last_10k_avg_distances = np.mean(last_10k_distances)  # Get Performance of Benchmark
    count = 1000
    rev_value = None
    for d in avg_distances_per_thousand_episodes:
        # Write for Every 1000 Episodes the Distance to Excel
        worksheet.cell(row=row_position + 3, column=7, value=count)
        worksheet.cell(row=row_position + 3, column=8, value=sum(d / 1000))
        # Write Relative Change for Learning every 1k Episodes
        if rev_value is None:
            ref_value = worksheet.cell(row=4, column=8).value
        worksheet.cell(row=row_position + 3, column=9, value=sum(d / 1000) / ref_value)
        count += 1000
        row_position += 1
    """Print and Write Result of Benchmark"""
    worksheet.cell(row=22, column=11, value=last_10k_avg_distances)
    elapsed_time = timeit.default_timer() - start_time
    worksheet.cell(row=4, column=11, value=elapsed_time)  # Write Computational Time
    worksheet.cell(row=25, column=11, value=vehicle.failure_result)
    workbook.save(workbook_name)  # Save and Close Excel File
    return last_10k_avg_distances

def main(instance):
    data, demand_bottom, demand_top, capacity, apriori_list, demand_mean = LoadIn_instance(instance)
    workbook, worksheet, dataset_name, workbook_name = saving_process(instance)
    vehicle = Service(0, capacity, 0, 0)
    for episode in range(num_episodes):
        vehicle.reset_variables(capacity, instance, apriori_list)
        for x in apriori_list:
            vehicle.execute_episode(demand_mean, apriori_list, data, capacity, worksheet, episode, x)
        vehicle.post_episode_calculation()
    last_10k_avg_distances = print_final(1, worksheet, workbook, workbook_name, vehicle)
    return last_10k_avg_distances




if __name__ == "__main__":
    instance = Instance('C109', 90, 70, 10, 50)
    last_10k_avg_distances = main(instance)
    print(last_10k_avg_distances)

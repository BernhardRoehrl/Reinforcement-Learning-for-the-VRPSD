import os
from openpyxl.styles import Alignment
from Instance_Prerequisites import Instance
import Reinforcement_Learning_Efficient
import Simple_Policy_Efficient
from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


"""Fixed"""
data = []  # for our dataframe
solomon_dir = r'D:\Code\VRP OR\Solomon Instances'  # location of Solomon Instances
solomon_files = [f for f in os.listdir(solomon_dir) if os.path.isfile(os.path.join(solomon_dir, f))]
# all solomon_files to be tested
"""Instance-Profiles"""
customer_sizes = [5, 10, 15, 20, 30, 40, 50, 60, 70, 80, 90, 100]   # all customer sizes to be tested
demand_configurations = [  # all demand_capacity_configurations to test
    {"demand_bottom": 1, "demand_top": 10, "capacity": 15},  # Small Demand, Small Surplus
    {"demand_bottom": 1, "demand_top": 10, "capacity": 100},  # Small Demand, Large Surplus
    {"demand_bottom": 10, "demand_top": 50, "capacity": 70},  # Moderate Demand, Moderate Surplus
    {"demand_bottom": 1, "demand_top": 99, "capacity": 100},  # Large Demand, Small Surplus
]
"""Progress Bar"""
total_iterations = len(solomon_files) * len(customer_sizes) * len(demand_configurations)
pbar = tqdm(total=total_iterations)



"""Core Loop"""
for solomon_file in solomon_files:  # Go through all Instances
    file_name = solomon_file
    for customer_size in customer_sizes:  # Go through all customer-sizes
        size = customer_size
        for demand_configuration in demand_configurations:  # Go through all configs
            demand_top = demand_configuration["demand_top"]
            demand_bottom = demand_configuration["demand_bottom"]
            capacity = demand_configuration["capacity"]
            instance = Instance(file_name, size, capacity, demand_bottom, demand_top)  # Create the unique instance
            result_SP, time_SP, failure_result_SP = Simple_Policy_Efficient.main(instance)  # Run instance in Simple Policy
            result_RL, time_RL, failure_result_RL = Reinforcement_Learning_Efficient.main(instance)  # Run instance in RL
            result_DF = result_SP - result_RL  # Get diff in Performance
            time_DF = time_SP - time_RL  # Get diff in Time
            failure_result_DF = failure_result_RL - failure_result_SP
            rl_result_imprv = ((result_SP - result_RL) / result_SP)
            demand_range = instance.demand_top - instance.demand_bottom
            capacity_surplus = instance.capacity - instance.demand_top
            capacity_ratio = demand_range / instance.capacity
            row = {  # Gather the data
                "Instance Name": instance.instance_name,
                "Instance ID": instance.profile["solomon_id"],
                "Customer Spread": instance.profile["customer_spread"],
                "Customer Size": instance.profile["N"],
                "Instance Size Category": instance.profile["InstanceSize"],
                "Capacity": instance.profile["capacity"],
                "Capacity Stress": instance.profile["Capacity Stress"],
                "Capacity Ratio": capacity_ratio,
                "Capacity Surplus": capacity_surplus,
                "Demand Lower Bound": instance.profile["Demand_Bottom"],
                "Demand Upper Bound": instance.profile["Demand_Top"],
                "Demand Range": demand_range,
                "RL Result": result_RL,
                "SP Result": result_SP,
                "Result-Deviation": result_DF,
                "RL Improvement %": rl_result_imprv,
                "Time SP in (s)": time_SP,
                "Time RL in (s)": time_RL,
                "Time DF in (s)": time_DF,
                "Failures SP": failure_result_SP,
                "Failures RL": failure_result_RL,
                "Failures Deviation": failure_result_DF,
            }
            data.append(row)  # Append the dictionary to the data list
            pbar.update()  # Update Progress Bar

"""Making Excel Look Good"""
# Write the DataFrame to Excel
df = pd.DataFrame(data)
df.to_excel("results.xlsx", index=False)
wb = openpyxl.load_workbook('results.xlsx')

# Chose Worksheet
ws = wb.active

# Table for Data Rows
tab = Table(displayName="Table1", ref=f"A1:V{ws.max_row}")

# Default Table Style
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

# Apply Table Style
tab.tableStyleInfo = style

# apply Table to Worksheet
ws.add_table(tab)

# Change sizing of columns for more readability
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 15)  # Determine how much room each column gets
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Center the values for more readability
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# Convert Columns to %
for cell in ws['P']:
    cell.number_format = '0.00%'
for cell in ws['H']:
    cell.number_format = '0.00%'
# save the changes
wb.save('results.xlsx')


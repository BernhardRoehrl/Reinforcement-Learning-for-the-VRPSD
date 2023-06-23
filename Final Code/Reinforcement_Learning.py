import random
import numpy as np
import pandas as pd
import timeit
import openpyxl
from Apriori import apriori_list
from Prerequisites import distance_matrix, capacity, demand_top, demand_bottom, customer_spread

"""Import Data"""
apriori_list = apriori_list  # import apriori_list
start_time = timeit.default_timer()  # Set start timer
data = {}  # Stores the data for the problem
data['distance_matrix'] = distance_matrix  # Import distance_matrix

"""Create Lists and Parameters"""
customer_list = []  # Init customer list
avg_distance = []  # Init avg_distance
last_10k_avg_distance = []  # Init List for Results out of Benchmarking
capacity = capacity  # Vehicle Capacity Parameter
demand_bottom = demand_bottom  # Smallest Value Demand customer can have
demand_top = demand_top  # Highest Value Demand customer can have

"""Additional Parameters for QLearning"""
action_refill = 0
action_serve = 1
action_space_size = [action_refill, action_serve]

"""Static Learning Parameters"""
# Exploration & Exploitation parameters
exploration_rate = 1.0  # Init Exploration rate
max_exploration_rate = 1.0  # Exploration probability at start
min_exploration_rate = 0.1  # Minimum exploration probability guaranteed outside of Benchmarking
rewards_all_episodes = []  # List of rewards
exploration_decay_rate = 0.00001  # Determines How fast to exploit more
q_table = np.zeros((capacity + 1, len(apriori_list), len(action_space_size)))  # init the q_table
discount_rate = 0.85  # Discounting rate (How much to value future Rewards)
learning_rate = 0.08  # How fast to learn --> determines episodes needed

"""Dynamic Learning Parameters"""
num_episodes = 30000+(capacity*(len(apriori_list)-2)*8)   # Dynamic Episode Sizing
num_episodes = int(np.ceil(num_episodes / 1000) * 1000)  # Round up num_episodes to a divider of 1000

def scale_hyperparameters(num_episodes):
    """Dynamic hyperparameters for learning based on the number of episodes. The learning_rate is calculated using a
    logarithmic function with a cube root. The exploration_decay_rate is calculated using an exponential function
    with a square root. Adjust coefficients as needed or just set values manually + and remove this function call"""
    learning_rate = -0.055 + 0.035 * np.log1p(np.cbrt(num_episodes - 30000))
    exploration_decay_rate = (0.21 ** np.log1p(np.sqrt(num_episodes - 30000))) * (1 / 10)
    return learning_rate, exploration_decay_rate

#learning_rate, exploration_decay_rate = scale_hyperparameters(num_episodes)

"""Saving Process"""
index = 1  # for looping later on
solution_prefix = "RL_"  # Indication of method
customer_spread = customer_spread  # C = Clustered or R = Random Spread Customers
dataset_name = customer_spread + "_N" + str(len(apriori_list) - 2) + "_H" + str(
    capacity) + "_D" + str(
    demand_bottom) + "-" + str(demand_top)  # Dynamic Instance Naming
print("\n3: Executing Reinforcement_Learning.py \nInstance: ", dataset_name)
workbook_name = f"{solution_prefix}Results.xlsx"
try:  # Excel File Handling
    workbook = openpyxl.load_workbook(workbook_name)
except FileNotFoundError:  # if Excel does not exist, create one
    workbook = openpyxl.Workbook()
    workbook.save(workbook_name)
if dataset_name in workbook.sheetnames:  # if Tab in Excel exists delete and recreate
    del workbook[dataset_name]
    worksheet = workbook.create_sheet(dataset_name)
else:  # Create if missing
    worksheet = workbook.create_sheet(dataset_name)
# Write Headers and existing values
worksheet.cell(row=1, column=1, value=dataset_name)
worksheet.cell(row=3, column=1, value='Customer')
worksheet.cell(row=3, column=2, value='Action')
worksheet.cell(row=3, column=3, value='Refill_Counter')
worksheet.cell(row=3, column=4, value='Failure_Counter')
worksheet.cell(row=3, column=5, value='Capacity at Decision')
worksheet.cell(row=3, column=7, value='Per Thousand Episodes')
worksheet.cell(row=3, column=8, value='Average Distance')
worksheet.cell(row=3, column=9, value='Relative Change')
worksheet.cell(row=3, column=11, value='Time for Computation in (s)')
worksheet.cell(row=6, column=11, value='Episodes')
worksheet.cell(row=7, column=11, value=num_episodes)
worksheet.cell(row=9, column=11, value='Learning Rate')
worksheet.cell(row=10, column=11, value=learning_rate)
worksheet.cell(row=12, column=11, value='Discount Rate')
worksheet.cell(row=13, column=11, value=discount_rate)
worksheet.cell(row=15, column=11, value='Min Exploration Rate')
worksheet.cell(row=16, column=11, value=min_exploration_rate)
worksheet.cell(row=18, column=11, value='Exploration Decay Rate')
worksheet.cell(row=19, column=11, value=exploration_decay_rate)
worksheet.cell(row=21, column=11, value='Result last 10k')
worksheet.cell(row=24, column=11, value='Failures')

"""Function & Classes"""


def create_customer():
    """Creates Customer_List with random demands & position from Apriori_List"""
    customer_list.clear()
    for y in apriori_list:
        customer_x = Customer(random.randrange(demand_bottom, demand_top, 1), y)
        if customer_x.position == 0:
            customer_x.demand = 0
        customer_list.append(customer_x)
    if capacity < demand_top:
        #  Forbidden due to how customer demand and route failure are calculated
        raise Exception("!!!Demand > Capacity!!!")
    return customer_list


class Service:
    """For the vehicle and tracking distance"""

    def __init__(self, position, capacity, distance, step_distance):
        self.exploration_rate = exploration_rate
        self.capacity = capacity
        self.step = 0
        self.old_capacity = 0  # Needed for Bellman Equation Calculations
        self.rewards_current_episodes = 0  # Track rewards of the current episode
        self.position = position
        self.distance = distance  # Track total distance throughout one episode
        self.step_distance = step_distance  # Track total distance for one customer
        self.refill_counter = 0  # For last episode tracking
        self.failure_counter = 0  # For last episode tracking
        self.failure_result = 0  # For whole Benchmarking Process

    def reset_variables(self, capacity):
        """Resetting for outer loop"""
        create_customer()
        self.position = 0
        self.capacity = capacity
        self.distance = 0
        self.step_distance = 0
        self.step = 0
        self.old_capacity = 0
        self.rewards_current_episodes = 0
        self.refill_counter = 0
        self.failure_counter = 0
        return self, customer_list

    def position_update(self, next_customer):
        """Update vehicle position"""
        self.position = next_customer
        return self.position

    def distance_update(self, target):
        """Update total vehicle distances travelled"""
        self.distance += data['distance_matrix'][self.position][target]  # Update total Episode Distance
        self.step_distance += data['distance_matrix'][self.position][target]  # Update Customer/Step Specific Distance
        return self.distance, self.step_distance

    def refill(self, current_step, customer):
        """The Agent chooses to refill. This means driving back and to the new customer"""
        """Move Vehicle from current position to the depot and refill"""
        self.distance_update(0)
        self.position_update(0)
        self.capacity = capacity  # Refilling the Capacity
        """Move Vehicle from depot to next customer and serve"""
        self.distance_update(customer)
        self.position_update(customer)
        self.capacity = self.capacity - customer_list[current_step].demand  # Update capacity
        customer_list[current_step].demand = 0  # Set customer demand to 0
        if episode > num_episodes - 10000:  # For Benchmarking statistics
            self.refill_counter += 1
        return self.position, self.capacity, self.distance, customer_list[
            current_step].demand, self.step_distance, self.refill_counter

    def serve(self, current_step, customer):
        """Function handling both scenarios for serving. Failure or successful 1st time-servings"""
        if self.capacity < customer_list[current_step].demand:
            """Failure: serve partially -> Go To depot -> Go back -> Fulfill serving"""
            # 1st Distance: Serve customer partially
            if episode > num_episodes - 10000:  # For statistics
                self.failure_counter += 1  # For Last Episode Overview
                self.failure_result += 1  # For Benchmarking statistics
            self.distance_update(customer)
            self.position_update(customer)
            customer_list[current_step].demand = customer_list[current_step].demand - self.capacity
            self.capacity = 0
            # 2nd Distance: To Depot from location of partially served customer
            self.distance_update(0)
            self.position_update(0)
            self.capacity = capacity
            # 3rd Distance: Back To Customer from Depot serving left over demand
            self.distance_update(customer)
            self.position_update(customer)
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            return self.position, self.failure_counter, self.capacity, self.distance, customer_list[
                current_step].demand, self.step_distance, self.failure_result
        else:
            """Vehicle can serve fully. Regular workflow"""
            self.capacity = self.capacity - customer_list[current_step].demand
            customer_list[current_step].demand = 0
            self.distance_update(customer)
            self.position_update(customer)
            return customer_list[current_step].demand, self.capacity, self.position, self.distance, self.step_distance

    def execute_episode(self):
        """Core function executing the service for one step/customer/decision epoch"""
        self.step_distance = 0  # Reset step distance at the beginning of each
        self.old_capacity = self.capacity  # Update Capacities
        if episode <= num_episodes - 10000:
            """Use E-Greedy if not Benchmarking"""
            exploration_rate_threshold = random.uniform(0, 1)  # Check against decaying exploration rate
            if exploration_rate_threshold < self.exploration_rate:
                """Explore"""
                action = random.choice(action_space_size)  # Choose Action on Random
                if action == 1:  # Agent chooses action 1: Vehicle must serve
                    self.serve(self.step, apriori_list[self.step])
                else:  # Agent chooses action 2: refilling
                    self.refill(self.step, apriori_list[self.step])
            else:
                """Exploit"""  # Agent chooses lowest Q-Value for state action combination
                action = np.argmin(q_table[self.old_capacity, apriori_list[self.step], :])
                if action == 1:  # Agent chooses action 1: Vehicle must serve
                    self.serve(self.step, apriori_list[self.step])
                elif action == 0:  # Agent chooses action 2: refilling
                    self.refill(self.step, apriori_list[self.step])
        else:
            """Dont use E-Greedy --> Benchmarking --> Exploit"""
            # Agent chooses lowest Q-Value for state action combination
            action = np.argmin(q_table[self.old_capacity, apriori_list[self.step], :])
            if action == 1:  # Agent chooses action 1: Vehicle must serve
                self.serve(self.step, apriori_list[self.step])
            elif action == 0:  # Agent chooses action 2: refilling
                self.refill(self.step, apriori_list[self.step])
        reward = self.step_distance  # Set immediate reward to current step_distance
        if episode < num_episodes - 10000:
            """If not Benchmarking --> Update Q-Table"""
            old_value = q_table[self.old_capacity, apriori_list[self.step], action]  # Set old Q(s,a)
            if self.step + 1 < len(apriori_list):  # Set best new Q'(s,a)
                best_expected_value = np.min(q_table[self.capacity, apriori_list[self.step + 1], :])
            else:  # Condition to not go out of bounds
                best_expected_value = 0
            bellman_term = (reward + discount_rate * best_expected_value - old_value)
            q_table[self.old_capacity, apriori_list[self.step], action] = old_value + learning_rate * bellman_term
        """Preparing next Cycle"""
        self.rewards_current_episodes += reward  # Accumulate step distance for whole episode
        self.step = self.step + 1  # Adjustments for next episode
        if episode == num_episodes - 1:  # Write last Episode to Excel for quick overview
            self.write_final_episode(action)
        return self, q_table

    def post_episode_calculation(self):
        """After each episode this updates the exploration rate according to e-greedy and
        adds the data for rewards and distances to the lists for later calculations"""
        avg_distance.append(self.distance)
        self.exploration_rate = min_exploration_rate + (max_exploration_rate - min_exploration_rate) * np.exp(
            -exploration_decay_rate * episode)  # decay exploration rate
        rewards_all_episodes.append(self.rewards_current_episodes)  # List for Learning analysis
        return avg_distance, self, rewards_all_episodes,

    def write_final_episode(self, action):
        """Function that writes the best routing found after training to text and excel"""
        worksheet.cell(row=self.step + 3, column=1, value=x)  # x = which customer
        worksheet.cell(row=self.step + 3, column=2, value=action)
        worksheet.cell(row=self.step + 3, column=3, value=self.refill_counter)
        worksheet.cell(row=self.step + 3, column=4, value=self.failure_counter)
        worksheet.cell(row=self.step + 3, column=5, value=self.old_capacity)


class Customer:
    """Customers with demand and position"""

    def __init__(self, demand, position):
        self.demand = demand
        self.position = position

    def set_demand(self, demand):
        """Change Customer Demand"""
        self.demand = demand


def print_final(row_position, q_table):
    """Function that incorporates all output related information for further usage, give index for row_position"""
    """Calculate All Kinds of Distances for Evaluation out of Lists"""
    avg_distances_per_thousand_episodes = np.array_split(np.array(avg_distance), num_episodes // 1000)
    slice_index = max(0, len(avg_distance) - 10000)
    last_10k_distances = avg_distance[slice_index:]
    last_10k_avg_distances = np.mean(last_10k_distances)  # Get Performance of Benchmark
    count = 1000
    rev_value = None
    for d in avg_distances_per_thousand_episodes:
        # Write for Every 1000 Episodes the Distance to Excel to keep Track of Learning
        worksheet.cell(row=row_position + 3, column=7, value=count)
        worksheet.cell(row=row_position + 3, column=8, value=sum(d / 1000))
        # Write Relative Change for Learning every 1k Episodes
        if rev_value is None:
            ref_value = worksheet.cell(row=4, column=8).value
        worksheet.cell(row=row_position + 3, column=9, value=sum(d / 1000) / ref_value)
        count += 1000
        row_position += 1
    """Print and Write Result of Benchmark"""
    worksheet.cell(row=22, column=11, value=last_10k_avg_distances)
    print('\n---> Result for 10k Avg= ', last_10k_avg_distances, "<----")
    worksheet.cell(row=4, column=11, value=elapsed_time)  # Write Computational Time
    worksheet.cell(row=25, column=11, value=vehicle.failure_result)
    """Processing the 3d Numpy Array: Q_table for Excel"""
    worksheet.cell(row=5, column=16, value='Q-Table')  # Header for q_table
    names = ['x', 'y', 'z']
    header_names = pd.MultiIndex.from_product([range(s) for s in q_table.shape], names=names)
    df = pd.DataFrame({'q_table': q_table.flatten()}, index=header_names)['q_table']
    df = df.unstack(level='z').swaplevel().sort_index()
    df.columns = ['Refill', 'Serve']
    df.index.names = ['Customer', 'Capacity']
    workbook.save(workbook_name)  # Save and Close Excel File for Pandas
    writer = pd.ExcelWriter(workbook_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    with writer as writer:
        df.to_excel(writer, sheet_name=dataset_name, startrow=5, startcol=15)


if __name__ == "__main__":
    vehicle = Service(0, capacity, 0, 0)  # Init the vehicle
    for episode in range(num_episodes):  # Outer Loop for num_episodes
        vehicle.reset_variables(capacity)  # Resetting the variables that got changed in inner loop
        for x in apriori_list:  # Inner Loop to let the vehicle go through all customers
            vehicle.execute_episode()  # Main function executing the service
        vehicle.post_episode_calculation()  # Capture and prepare calculations after every episode
    elapsed_time = timeit.default_timer() - start_time  # Computational Time
    print_final(index, q_table)  # Save results for later use
    print("learning_rate = ", learning_rate, "exploration_decay_rate = ", f'{exploration_decay_rate:.20f}')